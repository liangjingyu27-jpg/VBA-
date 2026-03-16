from dataclasses import dataclass
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.core.services.settlement_terms_service import SETTLEMENT_TERM_TEMPLATE_FIELDS, to_template_row


@dataclass
class ExcelLoadResult:
    workbook: Workbook
    source_sheet_name: str
    source_sheet_reason: str
    available_sheet_names: list[str]
    source_rows: int
    source_cols: int
    preview_rows: list[list[str]]


def safe_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def to_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if text == "":
        return 0.0
    try:
        return float(text)
    except Exception:
        return 0.0


def normalize_sheet_name(name: str) -> str:
    return str(name).strip().replace(" ", "")


def sheet_has_meaningful_content(sheet: Worksheet) -> bool:
    non_empty_count = 0
    for row in sheet.iter_rows(
        min_row=1,
        max_row=min(sheet.max_row, 20),
        max_col=min(sheet.max_column, 12),
        values_only=True,
    ):
        for value in row:
            if value is not None and str(value).strip() != "":
                non_empty_count += 1
                if non_empty_count >= 3:
                    return True
    return False


def select_source_sheet(workbook: Workbook) -> tuple[str, Worksheet | None, str]:
    sheets: list[Worksheet] = list(workbook.worksheets)
    if not sheets:
        return "尚未识别", None, "工作簿中未找到任何工作表"

    for sheet in sheets:
        if normalize_sheet_name(sheet.title) == "原始数据":
            return sheet.title, sheet, "精确命中《原始数据》"

    for sheet in sheets:
        if "原始数据" in normalize_sheet_name(sheet.title):
            return sheet.title, sheet, "名称包含“原始数据”"

    for sheet in sheets:
        if sheet_has_meaningful_content(sheet):
            return sheet.title, sheet, "未命中“原始数据”，兜底选择第一个有内容的工作表"

    first_sheet = sheets[0]
    return first_sheet.title, first_sheet, "未命中“原始数据”且未识别到明显内容，兜底选择第一个工作表"


def preview_rows_from_sheet(sheet: Worksheet, max_rows: int = 5, max_cols: int = 6) -> list[list[str]]:
    preview_rows: list[list[str]] = []

    for row in sheet.iter_rows(
        min_row=1,
        max_row=min(sheet.max_row, 12),
        max_col=min(sheet.max_column, max_cols),
        values_only=True,
    ):
        row_values = [safe_text(value) for value in row]
        if any(cell != "" for cell in row_values):
            preview_rows.append(row_values)

        if len(preview_rows) >= max_rows:
            break

    return preview_rows


def open_excel_with_source(file_path: str) -> ExcelLoadResult:
    workbook = load_workbook(file_path, data_only=True)

    selected_name, selected_sheet, reason = select_source_sheet(workbook)

    if selected_sheet is not None:
        source_rows = int(selected_sheet.max_row or 0)
        source_cols = int(selected_sheet.max_column or 0)
        preview_rows = preview_rows_from_sheet(selected_sheet)
    else:
        source_rows = 0
        source_cols = 0
        preview_rows = []

    return ExcelLoadResult(
        workbook=workbook,
        source_sheet_name=selected_name,
        source_sheet_reason=reason,
        available_sheet_names=workbook.sheetnames,
        source_rows=source_rows,
        source_cols=source_cols,
        preview_rows=preview_rows,
    )


def close_workbook_safely(workbook: Workbook | None) -> None:
    if workbook is None:
        return
    try:
        workbook.close()
    except Exception:
        pass


def find_sheet_by_keywords(workbook: Workbook, keywords: list[str]) -> Worksheet | None:
    normalized_keywords = [normalize_sheet_name(word) for word in keywords if normalize_sheet_name(word) != ""]
    for sheet in workbook.worksheets:
        normalized_title = normalize_sheet_name(sheet.title)
        for keyword in normalized_keywords:
            if keyword in normalized_title:
                return sheet
    return None


def sheet_to_dict_rows(sheet: Worksheet, header_row: int = 1) -> list[dict[str, Any]]:
    if sheet is None:
        return []

    header_cells = next(
        sheet.iter_rows(
            min_row=header_row,
            max_row=header_row,
            values_only=True,
        ),
        None,
    )
    if header_cells is None:
        return []

    headers: list[str] = []
    used = {}
    for idx, value in enumerate(header_cells, start=1):
        raw = safe_text(value)
        if raw == "":
            raw = f"未命名列_{idx}"
        if raw in used:
            used[raw] += 1
            raw = f"{raw}_{used[raw]}"
        else:
            used[raw] = 1
        headers.append(raw)

    dict_rows: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if row is None:
            continue
        if not any(value is not None and str(value).strip() != "" for value in row):
            continue

        row_dict: dict[str, Any] = {}
        for col_index, header in enumerate(headers):
            row_dict[header] = row[col_index] if col_index < len(row) else None
        dict_rows.append(row_dict)

    return dict_rows


def upsert_settlement_term_result(
    result_store: dict[str, dict[str, Any]],
    task_id: str,
    payload: dict[str, Any],
) -> bool:
    if task_id.strip() == "":
        return False

    template_row = to_template_row(payload)
    result_store[task_id] = {field: template_row.get(field, "") for field in SETTLEMENT_TERM_TEMPLATE_FIELDS}
    return True
